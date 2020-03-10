from django.shortcuts import render, redirect
from django.contrib.auth.models import User, auth
from django.contrib import messages
from .models import Ind_User, Org_User
from django.core.mail import EmailMessage
from django.core.files.storage import FileSystemStorage, default_storage
import os
from django.conf import settings
from django.http import HttpResponse, Http404
from openpyxl import Workbook
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
import reportlab
from reportlab.platypus import Image, Paragraph, Table, TableStyle, PageBreak, SimpleDocTemplate
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart, BarChartProperties, HorizontalBarChart

from reportlab.lib.colors import black, red, purple, green, maroon, brown, pink, white, HexColor, darkgray, PCMYKColor
from reportlab.graphics.charts.legends import Legend
import time
import datetime


# Splash Screen (First Page)

def splash(request):
    if request.user.is_authenticated:
        u = request.user.username
        u = str(u)
        if Ind_User.objects.filter(username=u).exists():
            u = request.user.username
            u = str(u)
            fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            values = []
            total = []
            values1 = []
            m1 = datetime.date.today().month
            m1 = int(m1)
            for i in range(1, 12):
                ref = ws.cell(row=m1 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values.append(ref_value)

            for i in range(1, 12):
                ref = ws1.cell(row=m1 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values1.append(ref_value)

            labels = ['loans', 'utility bills', 'insurance', 'entertainment', 'groceries', 'transportation',
                      'retirement fund',
                      'emergency fund', 'childcare', 'clothing', 'maintenance']
            # fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
            # print(fn)
            # workbook = load_workbook(fn)
            # ws = workbook.get_sheet_by_name('budget')
            # ws = workbook.active
            # print(a1.value)

            # Fetching the category and amount

            if request.method == 'POST':
                m = request.POST['mymonth']
                m = str(m)
                m = m.lower()
                # print(m)

                if m == 'none':
                    messages.add_message(request, messages.INFO, 'Error No Month Selected')
                if m == 'january':
                    s = 1
                if m == 'february':
                    s = 2
                if m == 'march':
                    s = 3
                if m == 'april':
                    s = 4
                if m == 'may':
                    s = 5
                if m == 'june':
                    s = 6
                if m == 'july':
                    s = 7
                if m == 'august':
                    s = 8
                if m == 'september':
                    s = 9
                if m == 'october':
                    s = 10
                if m == 'november':
                    s = 11
                if m == 'december':
                    s = 12
                l = request.POST['myselection']
                l = str(l)
                l = l.lower()
                # print(l)
                bal = request.POST['exp']
                # print(bal)
                row = 1
                column = 1

                # Adding Data to Excel

                for i in range(1, 12):
                    ref = ws.cell(row=row, column=i)
                    ref_value = ref.value
                    ref_value = str(ref_value)
                    ref_value = ref_value.lower()
                    # print(ref_value)
                    if ref_value == l:
                        ws.cell(row=row + 1, column=i, value=int(bal))
                        workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                        break
                    else:
                        continue

        elif Org_User.objects.filter(username=u).exists():
            fn = (settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_Data_1.xlsx')
            print(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('budget')
            # ws = workbook.active
            # print(a1.value)

            # Fetching the category and amount

            if request.method == 'POST':
                m = request.POST['mymonth1']
                m = str(m)
                m = m.lower()
                # print(m)
                if m == 'none':
                    messages.add_message(request, messages.INFO, 'Error No Month Selected')
                if m == 'january':
                    s = 1
                if m == 'february':
                    s = 2
                if m == 'march':
                    s = 3
                if m == 'april':
                    s = 4
                if m == 'may':
                    s = 5
                if m == 'june':
                    s = 6
                if m == 'july':
                    s = 7
                if m == 'august':
                    s = 8
                if m == 'september':
                    s = 9
                if m == 'october':
                    s = 10
                if m == 'november':
                    s = 11
                if m == 'december':
                    s = 12
                l = request.POST['myselection']
                l = str(l)
                l = l.lower()
                # print(l)
                bal = request.POST['exp']
                # print(bal)
                row = 1
                column = 1

                # Adding Data to Excel

                for i in range(1, 12):
                    ref = ws.cell(row=row, column=i)
                    ref_value = ref.value
                    ref_value = str(ref_value)
                    ref_value = ref_value.lower()
                    # print(ref_value)
                    if ref_value == l:
                        ws.cell(row=row + 1, column=i, value=int(bal))
                        workbook.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_Data_1.xlsx')
                        break
                    else:
                        continue

    # # Calculating Total Column
    #             for i in range(2,13):
    #                 total = 0
    #                 for j in range(2,12):
    #                     ref = ws.cell(row = row, column = i)
    #                     ref_value = ref.value
    #                     ref_value = (ref_value)
    #                     total = total + ref_value
    #                     ws.cell(row = i, column = 12, value = total)

    return render(request, 'accounts/splash.html')


# Individual User Register and excel creation

def ind_register_view(request):
    if request.method == 'POST':
        f_name = request.POST['f_name']
        l_name = request.POST['l_name']
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        if password1 == password2:
            if User.objects.filter(username=username).exists():
                print("Username Exists")
            elif User.objects.filter(email=email).exists():
                print("email Taken")
            else:
                user = User.objects.create_user(username=username, password=password1, email=email, first_name=f_name,
                                                last_name=l_name)
                user.save();
                s = Ind_User(username=username, l_name=l_name, f_name=f_name, password=password1, email=email)
                s.save();
                headers = ['loans', 'utility bills', 'insurance', 'entertainment', 'groceries', 'transportation',
                           'retirement fund', 'emergency fund', 'childcare and school costs', 'clothing',
                           'maintainance', 'total']
                wb = Workbook()
                ws = wb.active
                ws.title = 'budget'
                ws.append(headers)
                ws1 = wb.create_sheet('expenses')
                ws1.append(headers)
                u = username
                u = str(u)
                wb.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                fn = (r'uploads\Indiv_' + u + '_Data.xlsx')
                s.e_file = fn
                s.save()
        else:
            print("Password does not match")
        return redirect('/')
    else:
        return render(request, 'accounts/register_ind.html')


# Individual User Login

def ind_login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', False)
        password1 = request.POST.get('password1', False)

        user = auth.authenticate(username=username, password=password1)

        if user is not None:
            auth.login(request, user)
            return redirect('../../accounts/home')

        else:
            print("Error")
            messages.info(request, 'Invalid Username or Password')
    return render(request, 'accounts/login_ind.html')


# For Budget
def home(request):
    # Checking current user and loading excel

    if request.user.is_authenticated:
        u = request.user.username
        u = str(u)
        # if u == 'hitesh25':
        #     fn = (settings.MEDIA_ROOT+r'\uploads\Indiv_'+u+'_Data.xlsx')
        # elif u == 'hitesh98':
        #     fn = (settings.MEDIA_ROOT+r'\uploads\Organ_'+u+'_Data_1.xlsx')
        # fn = str(fn)
        if Ind_User.objects.filter(username=u).exists():
            u = request.user.username
            u = str(u)
            fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            values = []
            total = []
            total1 = []
            values1 = []
            global m1
            m1 = datetime.date.today().month
            m1 = int(m1)
            for i in range(1, 12):
                ref = ws.cell(row=m1 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values.append(ref_value)

            for i in range(1, 12):
                ref = ws1.cell(row=m1 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values1.append(ref_value)

            for j in range(2, 14):
                ref = ws.cell(row=j, column=12)
                ref_value = ref.value
                if ref_value == None:
                    total.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    total.append(ref_value)

            for j in range(2, 14):
                ref = ws1.cell(row=j, column=12)
                ref_value = ref.value
                if ref_value == None:
                    total1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    total1.append(ref_value)

            if m1 == 1:
                m2 = 'January'
            if m1 == 2:
                m2 = 'February'
            if m1 == 3:
                m2 = 'March'
            if m1 == 4:
                m2 = 'April'
            if m1 == 5:
                m2 = 'May'
            if m1 == 6:
                m2 = 'June'
            if m1 == 7:
                m2 = 'July'
            if m1 == 8:
                m2 = 'August'
            if m1 == 9:
                m2 = 'September'
            if m1 == 10:
                m2 = 'October'
            if m1 == 11:
                m2 = 'November'
            if m1 == 12:
                m2 = 'December'
            val_total = total[m1 - 1]
            val_total1 = total1[m1 - 1]

            rem_budget = val_total1 - val_total

            hig = values[:]
            del (hig[-1])
            max_exp = max(hig)

            labels = ['loans', 'utility bills', 'insurance', 'entertainment', 'groceries', 'transportation',
                      'retirement fund',
                      'emergency fund', 'childcare', 'clothing', 'maintenance']
            context = {'values': values, 'values1': values1, 'month': m2, 'labels': labels, 'total': val_total,
                       'total1': val_total1, 'rem_budget': rem_budget, 'max_exp': max_exp}

            # Fetching the category and amount

            if request.method == 'POST':
                if request.method == 'POST' and 'Submit' in request.POST:
                    print('ok ok')
                    m = request.POST['mymonth']
                    m = str(m)
                    m = m.lower()
                    print(m)
                    if m == 'none':
                        messages.add_message(request, messages.INFO, 'Error No Month Selected')
                    if m == 'january':
                        s = 1
                    if m == 'february':
                        s = 2
                    if m == 'march':
                        s = 3
                    if m == 'april':
                        s = 4
                    if m == 'may':
                        s = 5
                    if m == 'june':
                        s = 6
                    if m == 'july':
                        s = 7
                    if m == 'august':
                        s = 8
                    if m == 'september':
                        s = 9
                    if m == 'october':
                        s = 10
                    if m == 'november':
                        s = 11
                    if m == 'december':
                        s = 12
                    l = request.POST['myselection']
                    l = str(l)
                    l = l.lower()
                    print(l)
                    bal = request.POST['exp']
                    print(bal)
                    print(s)
                    row = 1
                    column = 1

                    # Adding Data to Excel

                    for i in range(1, 12):
                        ref = ws.cell(row=row, column=i)
                        ref_value = ref.value
                        ref_value = str(ref_value)
                        ref_value = ref_value.lower()
                        # print(ref_value)
                        if ref_value == l:
                            ws.cell(row=s + 1, column=i, value=int(bal))
                            workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                            break
                        else:
                            continue

                elif request.method == 'POST' and 'Submit1' in request.POST:
                    print('yes yes')
                    m = request.POST['mymonth1']
                    m = str(m)
                    m = m.lower()
                    print(m)
                    if m == 'none':
                        messages.add_message(request, messages.INFO, 'Error No Month Selected')
                    if m == 'january':
                        s = 1
                    if m == 'february':
                        s = 2
                    if m == 'march':
                        s = 3
                    if m == 'april':
                        s = 4
                    if m == 'may':
                        s = 5
                    if m == 'june':
                        s = 6
                    if m == 'july':
                        s = 7
                    if m == 'august':
                        s = 8
                    if m == 'september':
                        s = 9
                    if m == 'october':
                        s = 10
                    if m == 'november':
                        s = 11
                    if m == 'december':
                        s = 12
                    l = request.POST['myselection1']
                    l = str(l)
                    l = l.lower()
                    print(l)
                    bal = request.POST['exp1']
                    print(bal)
                    print(s)
                    row = 1
                    column = 1

                    # Adding Data to Excel

                    for i in range(1, 12):
                        ref = ws1.cell(row=row, column=i)
                        ref_value = ref.value
                        ref_value = str(ref_value)
                        ref_value = ref_value.lower()
                        # print(ref_value)
                        if ref_value == l:
                            ws1.cell(row=s + 1, column=i, value=int(bal))
                            workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                            break
                        else:
                            continue

            # Calculating Total Column

            fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            for i in range(2,14):
                total12 = 0
                total13 = 0
                for j in range(1,12):
                    ref = ws.cell(row = i, column = j)
                    ref1 = ws1.cell(row=i, column = j)
                    ref_value1 = ref1.value
                    ref_value = ref.value
                    if ref_value == None:
                        ref_value = 0
                    if ref_value1 == None:
                        ref_value1 = 0
                    total12 = total12 + ref_value
                    total13 = total13 + ref_value1
                ws.cell(row = i, column = 12).value = total12
                ws1.cell(row = i, column = 12).value = total13
                workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')

            return render(request, 'accounts/home.html', context=context)
        return redirect('/accounts/home.html')


def home_corp(request):
    # Checking current user and loading excel
    global m10
    if request.user.is_authenticated:
        u = request.user.username
        u = str(u)
        if Org_User.objects.filter(username=u).exists():
            u = request.user.username
            u = str(u)
            fn = (settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
            dep = fn.split('_')
            cur_dep = dep[-1]
            cur_dep = cur_dep.split('.')
            print(cur_dep[0])
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            values = []
            total = []
            total1 = []
            values1 = []
            m10 = datetime.date.today().month
            m10 = int(m10)
            print(m10)
            for i in range(1, 11):
                ref = ws.cell(row=m10 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values.append(ref_value)

            for i in range(1, 11):
                ref = ws1.cell(row=m10 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values1.append(ref_value)

            for j in range(2, 14):
                ref = ws.cell(row=j, column=11)
                ref_value = ref.value
                if ref_value == None:
                    total.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    total.append(ref_value)

            for j in range(2, 14):
                ref = ws1.cell(row=j, column=11)
                ref_value = ref.value
                if ref_value == None:
                    total1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    total1.append(ref_value)

            if m10 == 1:
                m2 = 'January'
            if m10 == 2:
                m2 = 'February'
            if m10 == 3:
                m2 = 'March'
            if m10 == 4:
                m2 = 'April'
            if m10 == 5:
                m2 = 'May'
            if m10 == 6:
                m2 = 'June'
            if m10 == 7:
                m2 = 'July'
            if m10 == 8:
                m2 = 'August'
            if m10 == 9:
                m2 = 'September'
            if m10 == 10:
                m2 = 'October'
            if m10 == 11:
                m2 = 'November'
            if m10 == 12:
                m2 = 'December'

            val_total = total[m10 - 1]
            val_total1 = total1[m10 - 1]
            print(val_total1)

            rem_budget = val_total1 - val_total

            hig = values[:]
            del (hig[-1])
            max_exp = max(hig)

            labels = ['loans', 'salaries', 'maintenance', 'inventory', 'party fund', 'variable costs', 'bonuses',
                      'operation losses', 'travel expenses', 'charity', 'total']
            context = {'values': values, 'values1': values1, 'month': m2, 'labels': labels, 'total': val_total,
                       'total1': val_total1, 'rem_budget': rem_budget, 'max_exp': max_exp}

            if request.method == 'POST':
                if request.method == 'POST' and 'Submit' in request.POST:
                    print('ok ok')
                    m = request.POST['mymonth']
                    m = str(m)
                    m = m.lower()
                    print(m)
                    if m == 'none':
                        messages.add_message(request, messages.INFO, 'Error No Month Selected')
                    if m == 'january':
                        s = 1
                    if m == 'february':
                        s = 2
                    if m == 'march':
                        s = 3
                    if m == 'april':
                        s = 4
                    if m == 'may':
                        s = 5
                    if m == 'june':
                        s = 6
                    if m == 'july':
                        s = 7
                    if m == 'august':
                        s = 8
                    if m == 'september':
                        s = 9
                    if m == 'october':
                        s = 10
                    if m == 'november':
                        s = 11
                    if m == 'december':
                        s = 12
                    l = request.POST['myselection']
                    l = str(l)
                    l = l.lower()
                    print(l)
                    bal = request.POST['exp']
                    print(bal)
                    print(s)
                    row = 1
                    column = 1

                    # Adding Data to Excel

                    for i in range(1, 12):
                        ref = ws.cell(row=row, column=i)
                        ref_value = ref.value
                        ref_value = str(ref_value)
                        ref_value = ref_value.lower()
                        # print(ref_value)
                        if ref_value == l:
                            ws.cell(row=s + 1, column=i, value=int(bal))
                            workbook.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
                            break
                        else:
                            continue

                elif request.method == 'POST' and 'Submit1' in request.POST:
                    print('yes yes')
                    m = request.POST['mymonth1']
                    m = str(m)
                    m = m.lower()
                    print(m)
                    if m == 'none':
                        messages.add_message(request, messages.INFO, 'Error No Month Selected')
                    if m == 'january':
                        s = 1
                    if m == 'february':
                        s = 2
                    if m == 'march':
                        s = 3
                    if m == 'april':
                        s = 4
                    if m == 'may':
                        s = 5
                    if m == 'june':
                        s = 6
                    if m == 'july':
                        s = 7
                    if m == 'august':
                        s = 8
                    if m == 'september':
                        s = 9
                    if m == 'october':
                        s = 10
                    if m == 'november':
                        s = 11
                    if m == 'december':
                        s = 12
                    l = request.POST['myselection1']
                    l = str(l)
                    l = l.lower()
                    print(l)
                    bal = request.POST['exp1']
                    print(bal)
                    print(s)
                    row = 1
                    column = 1

                    # Adding Data to Excel

                    for i in range(1, 12):
                        ref = ws1.cell(row=row, column=i)
                        ref_value = ref.value
                        ref_value = str(ref_value)
                        ref_value = ref_value.lower()
                        # print(ref_value)
                        if ref_value == l:
                            ws1.cell(row=s + 1, column=i, value=int(bal))
                            workbook.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
                            break
                        else:
                            continue

            # Calculating Total Column

            fn = (settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            for i in range(2,14):
                total12 = 0
                total13 = 0
                for j in range(1,11):
                    ref = ws.cell(row = i, column = j)
                    ref1 = ws1.cell(row=i, column = j)
                    ref_value1 = ref1.value
                    ref_value = ref.value
                    if ref_value == None:
                        ref_value = 0
                    if ref_value1 == None:
                        ref_value1 = 0
                    total12 = total12 + ref_value
                    total13 = total13 + ref_value1
                ws.cell(row = i, column = 11).value = total12
                ws1.cell(row = i, column = 11).value = total13
                workbook.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
                # print(total12)
                # print(total13)
            return render(request, 'accounts/home_corp.html', context=context)
    return render(request, 'accounts/home_corp.html')


def home1(request, int_object):
    global m1
    m1 = int(int_object)
    if request.user.is_authenticated:
        u = request.user.username
        u = str(u)
        if Ind_User.objects.filter(username=u).exists():
            u = request.user.username
            u = str(u)
            fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            values = []
            total = []
            total1 = []
            values1 = []
            for i in range(1, 12):
                ref = ws.cell(row=m1 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values.append(ref_value)

            for i in range(1, 12):
                ref = ws1.cell(row=m1 + 1, column=i)
                ref_value = ref.value
                if ref_value == None:
                    values1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    values1.append(ref_value)

            for j in range(2, 14):
                ref = ws.cell(row=j, column=12)
                ref_value = ref.value
                if ref_value == None:
                    total.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    total.append(ref_value)

            for j in range(2, 14):
                ref = ws1.cell(row=j, column=12)
                ref_value = ref.value
                if ref_value == None:
                    total1.append(0)
                elif ref_value != None:
                    ref_value = int(ref_value)
                    total1.append(ref_value)

            if m1 == 1:
                m2 = 'January'
            if m1 == 2:
                m2 = 'February'
            if m1 == 3:
                m2 = 'March'
            if m1 == 4:
                m2 = 'April'
            if m1 == 5:
                m2 = 'May'
            if m1 == 6:
                m2 = 'June'
            if m1 == 7:
                m2 = 'July'
            if m1 == 8:
                m2 = 'August'
            if m1 == 9:
                m2 = 'September'
            if m1 == 10:
                m2 = 'October'
            if m1 == 11:
                m2 = 'November'
            if m1 == 12:
                m2 = 'December'
            val_total = total[m1 - 1]
            val_total1 = total1[m1 - 1]
            # print(total)
            # print(total1)
            rem_budget = val_total1 - val_total

            hig = values[:]
            del (hig[-1])
            max_exp = max(hig)

            labels = ['loans', 'utility bills', 'insurance', 'entertainment', 'groceries', 'transportation',
                      'retirement fund',
                      'emergency fund', 'childcare', 'clothing', 'maintenance']
            context = {'values': values, 'values1': values1, 'month': m2, 'labels': labels, 'total': val_total,
                       'total1': val_total1, 'rem_budget': rem_budget, 'max_exp': max_exp}

            if request.method == 'POST':
                # if request.POST.get('Submit') == 'Submit_Button':
                # if ('Submit' in request.POST) == False:
                # if 'mymonth' in request.POST:
                if request.method == 'POST' and 'Submit' in request.POST:
                    print('ok ok')
                    m = request.POST['mymonth']
                    m = str(m)
                    m = m.lower()
                    print(m)
                    if m == 'none':
                        messages.add_message(request, messages.INFO, 'Error No Month Selected')
                    if m == 'january':
                        s = 1
                    if m == 'february':
                        s = 2
                    if m == 'march':
                        s = 3
                    if m == 'april':
                        s = 4
                    if m == 'may':
                        s = 5
                    if m == 'june':
                        s = 6
                    if m == 'july':
                        s = 7
                    if m == 'august':
                        s = 8
                    if m == 'september':
                        s = 9
                    if m == 'october':
                        s = 10
                    if m == 'november':
                        s = 11
                    if m == 'december':
                        s = 12
                    l = request.POST['myselection']
                    l = str(l)
                    l = l.lower()
                    print(l)
                    bal = request.POST['exp']
                    print(bal)
                    print(s)
                    row = 1
                    column = 1

                    # Adding Data to Excel

                    for i in range(1, 12):
                        ref = ws.cell(row=row, column=i)
                        ref_value = ref.value
                        ref_value = str(ref_value)
                        ref_value = ref_value.lower()
                        # print(ref_value)
                        if ref_value == l:
                            ws.cell(row=s + 1, column=i, value=int(bal))
                            workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                            break
                        else:
                            continue

                # elif request.POST.get('Submit1') == 'Submit_Button_1':
                # elif ('Submit1' in request.POST) == False:
                # elif 'mymonth1' in request.POST:
                elif request.method == 'POST' and 'Submit1' in request.POST:
                    print('yes yes')
                    m = request.POST['mymonth1']
                    m = str(m)
                    m = m.lower()
                    print(m)
                    if m == 'none':
                        messages.add_message(request, messages.INFO, 'Error No Month Selected')
                    if m == 'january':
                        s = 1
                    if m == 'february':
                        s = 2
                    if m == 'march':
                        s = 3
                    if m == 'april':
                        s = 4
                    if m == 'may':
                        s = 5
                    if m == 'june':
                        s = 6
                    if m == 'july':
                        s = 7
                    if m == 'august':
                        s = 8
                    if m == 'september':
                        s = 9
                    if m == 'october':
                        s = 10
                    if m == 'november':
                        s = 11
                    if m == 'december':
                        s = 12
                    l = request.POST['myselection1']
                    l = str(l)
                    l = l.lower()
                    print(l)
                    bal = request.POST['exp1']
                    print(bal)
                    print(s)
                    row = 1
                    column = 1

                    # Adding Data to Excel

                    for i in range(1, 12):
                        ref = ws1.cell(row=row, column=i)
                        ref_value = ref.value
                        ref_value = str(ref_value)
                        ref_value = ref_value.lower()
                        # print(ref_value)
                        if ref_value == l:
                            ws1.cell(row=s + 1, column=i, value=int(bal))
                            workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                            break
                        else:
                            continue

            fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
            fn = str(fn)
            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            ws1 = workbook.get_sheet_by_name('budget')
            for i in range(2,14):
                total12 = 0
                total13 = 0
                for j in range(1,12):
                    ref = ws.cell(row = i, column = j)
                    ref1 = ws1.cell(row=i, column = j)
                    ref_value1 = ref1.value
                    ref_value = ref.value
                    if ref_value == None:
                        ref_value = 0
                    if ref_value1 == None:
                        ref_value1 = 0
                    total12 = total12 + ref_value
                    total13 = total13 + ref_value1
                ws.cell(row = i, column = 12).value = total12
                ws1.cell(row = i, column = 12).value = total13
                workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')

            return render(request, 'accounts/home.html', context=context)
        return redirect('/accounts/home.html')


# Individual User Logout

def logout_view(request):
    auth.logout(request)
    return redirect('/')


def graph_view(request):
    u = request.user.username
    u = str(u)
    fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
    fn = str(fn)
    workbook = load_workbook(fn)
    ws = workbook.get_sheet_by_name('expenses')
    values = []
    total = []

    for i in range(1, 12):
        ref = ws.cell(row=m1 + 1, column=i)
        ref_value = ref.value
        if ref_value == None:
            values.append(0)
        elif ref_value != None:
            ref_value = int(ref_value)
            values.append(ref_value)

    for j in range(2, 14):
        ref = ws.cell(row=j, column=12)
        ref_value = ref.value
        if ref_value == None:
            total.append(0)
        elif ref_value != None:
            ref_value = int(ref_value)
            total.append(ref_value)

    if m1 == 1:
        m2 = 'January'
    if m1 == 2:
        m2 = 'February'
    if m1 == 3:
        m2 = 'March'
    if m1 == 4:
        m2 = 'April'
    if m1 == 5:
        m2 = 'May'
    if m1 == 6:
        m2 = 'June'
    if m1 == 7:
        m2 = 'July'
    if m1 == 8:
        m2 = 'August'
    if m1 == 9:
        m2 = 'September'
    if m1 == 10:
        m2 = 'October'
    if m1 == 11:
        m2 = 'November'
    if m1 == 12:
        m2 = 'December'

        # print(total)
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'Ocotber',
              'November', 'December']
    labels = ['loans', 'utility bills', 'insurance', 'entertainment', 'groceries', 'transportation', 'retirement fund',
              'emergency fund', 'childcare and school costs', 'clothing', 'maintenance']

    # p = make_subplots(rows=2, cols=1)
    # plot_div1 = plot(go.Figure(data=[go.Pie(labels=labels, values=values)]))
    # plot_div2 = plot(go.Figure(data=[go.Bar(x=months, y=total)]))
    return render(request, "accounts/home_graph.html",
                  context={'months': months, 'labels': labels, 'values': values, 'total': total, 'month': m1,
                           'month_name': m2})

def graph_view_corp(request):
    u = request.user.username
    u = str(u)
    fn = (settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
    fn = str(fn)
    workbook = load_workbook(fn)
    ws = workbook.get_sheet_by_name('expenses')
    values = []
    total = []

    for i in range(1, 11):
        ref = ws.cell(row=m10 + 1, column=i)
        ref_value = ref.value
        if ref_value == None:
            values.append(0)
        elif ref_value != None:
            ref_value = int(ref_value)
            values.append(ref_value)

    for j in range(2, 14):
        ref = ws.cell(row=j, column=12)
        ref_value = ref.value
        if ref_value == None:
            total.append(0)
        elif ref_value != None:
            ref_value = int(ref_value)
            total.append(ref_value)

    if m10 == 1:
        m2 = 'January'
    if m10 == 2:
        m2 = 'February'
    if m10 == 3:
        m2 = 'March'
    if m10 == 4:
        m2 = 'April'
    if m10 == 5:
        m2 = 'May'
    if m10 == 6:
        m2 = 'June'
    if m10 == 7:
        m2 = 'July'
    if m10 == 8:
        m2 = 'August'
    if m10 == 9:
        m2 = 'September'
    if m10 == 10:
        m2 = 'October'
    if m10 == 11:
        m2 = 'November'
    if m10 == 12:
        m2 = 'December'
    print(m10)
    print(m2)
        # print(total)
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'Ocotber',
              'November', 'December']
    labels = ['loans', 'salaries', 'maintenance', 'inventory', 'party fund', 'variable costs', 'bonuses',
              'operation losses', 'travel expenses', 'charity', 'total']

    return render(request, "accounts/home_graph_corp.html",
                  context={'months': months, 'labels': labels, 'values': values, 'total': total, 'month': m10,
                           'month_name': m2})


# For Expenses
def expense_view(request):
    # Fetching current user name and loading excel

    if request.user.is_authenticated:
        u = request.user.username
        u = str(u)
        fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
        fn = str(fn)
        workbook = load_workbook(fn)
        ws = workbook.get_sheet_by_name('expenses')
        # ws = workbook.active
        # print(a1.value)
        if request.method == 'POST':

            # Fetching the month, category and amount

            m = request.POST['mymonth']
            m = str(m)
            m = m.lower()
            # print(m)
            if m == 'none':
                messages.add_message(request, messages.INFO, 'Error No Month Selected')
            if m == 'january':
                s = 1
            if m == 'february':
                s = 2
            if m == 'march':
                s = 3
            if m == 'april':
                s = 4
            if m == 'may':
                s = 5
            if m == 'june':
                s = 6
            if m == 'july':
                s = 7
            if m == 'august':
                s = 8
            if m == 'september':
                s = 9
            if m == 'october':
                s = 10
            if m == 'november':
                s = 11
            if m == 'december':
                s = 12
            # print(s)
            l = request.POST['myselection']
            l = str(l)
            l = l.lower()
            # print(l)
            bal = request.POST['exp']
            # print(bal)
            row = 1
            column = 1

            # Adding Data to Excel

            for i in range(1, 12):
                ref = ws.cell(row=row, column=i)
                ref_value = ref.value
                ref_value = str(ref_value)
                ref_value = ref_value.lower()
                # print(ref_value)
                if ref_value == l:
                    ws.cell(row=s + 1, column=i, value=int(bal))
                    workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')
                    break
                else:
                    continue

            # Calculating Total Column

            workbook = load_workbook(fn)
            ws = workbook.get_sheet_by_name('expenses')
            for i in range(2, 14):
                total = 0
                for j in range(1, 12):
                    ref = ws.cell(row=i, column=j)
                    ref_value = ref.value
                    if ref_value == None:
                        continue
                    else:
                        total = total + ref_value
                        ws.cell(row=i, column=12, value=total)

                workbook.save(settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')

    return render(request, 'accounts/home_expenses.html')


# Organisations Register and excel creation for the 4 Departments Selected

def org_register_view(request):
    if request.method == 'POST':
        c_name = request.POST['c_name']
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        dep = request.POST['departments']
        print(dep)
        if password1 == password2:
            if User.objects.filter(username=username).exists():
                print("Username Exists")
            elif User.objects.filter(email=email).exists():
                print("email Taken")
            else:
                user = User.objects.create_user(username=username, password=password1, email=email, first_name=c_name)
                user.save();
                s = Org_User(username=username, c_name=c_name, password=password1, email=email)
                s.save();
                headers = ['loans', 'salaries', 'maintenance', 'inventory', 'party fund', 'variable costs', 'bonuses',
                           'operation losses', 'travel expenses', 'charity', 'total']

                # Workbook 1 creation
                wb1 = Workbook()
                ws1 = wb1.active
                ws1.title = 'budget'
                ws1.append(headers)
                ws2 = wb1.create_sheet('expenses')
                ws2.append(headers)
                u = username
                u = str(u)
                wb1.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')
                fn = (r'uploads\Organ_' + u + '_General.xlsx')
                s.e_file = fn
                s.save()

                # Workbook 2 Creation
                wb2 = Workbook()
                ws1 = wb2.active
                ws1.title = 'budget'
                ws1.append(headers)
                ws2 = wb2.create_sheet('expenses')
                ws2.append(headers)
                u = username
                u = str(u)
                wb2.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_Marketing.xlsx')
                fn = (r'uploads\Organ_' + u + '_Marketing.xlsx')
                s.e_file = fn
                s.save()

                # Workbook 3 Creation
                wb3 = Workbook()
                ws1 = wb3.active
                ws1.title = 'budget'
                ws1.append(headers)
                ws2 = wb3.create_sheet('expenses')
                ws2.append(headers)
                u = username
                u = str(u)
                wb3.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_Sales.xlsx')
                fn = (r'uploads\Organ_' + u + '_Sales.xlsx')
                s.e_file = fn
                s.save()

                # Workbook 4 Creation
                wb4 = Workbook()
                ws1 = wb4.active
                ws1.title = 'budget'
                ws1.append(headers)
                ws2 = wb4.create_sheet('expenses')
                ws2.append(headers)
                u = username
                u = str(u)
                wb4.save(settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_HR.xlsx')
                fn = (r'uploads\Organ_' + u + '_HR.xlsx')
                s.e_file = fn
                s.save()
        else:
            print("Password does not match")
        return redirect('/')
    else:
        return render(request, 'accounts/register_corp.html')


# Organization Login

def org_login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password1 = request.POST['password1']

        user = auth.authenticate(username=username, password=password1)

        if user is not None:
            auth.login(request, user)
            return redirect('../../accounts/home_corp')

        else:
            print("Error")
            messages.info(request, 'Invalid Username or Password')
    return render(request, 'accounts/login_corp.html')


# PDF Creation and instant deletion

def indiv_pdf_view(request):
    u = request.user.username
    user = Ind_User.objects.get(username=u)
    f_name = user.f_name
    l_name = user.l_name
    email = user.email
    name = f_name + ' ' + l_name
    saving = settings.MEDIA_ROOT + r'\uploads\Report_for_' + u + '.pdf'
    c = canvas.Canvas(saving, pagesize=A4)
    if m1 == 1:
        m2 = 'January'
    if m1 == 2:
        m2 = 'February'
    if m1 == 3:
        m2 = 'March'
    if m1 == 4:
        m2 = 'April'
    if m1 == 5:
        m2 = 'May'
    if m1 == 6:
        m2 = 'June'
    if m1 == 7:
        m2 = 'July'
    if m1 == 8:
        m2 = 'August'
    if m1 == 9:
        m2 = 'September'
    if m1 == 10:
        m2 = 'October'
    if m1 == 11:
        m2 = 'November'
    if m1 == 12:
        m2 = 'December'

    c.setFont('Times-Roman', 30)
    c.setFillColor(colors.deepskyblue)
    c.drawString(40, 780, 'Budget Tracker')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 13)
    c.drawString(55, 760, 'Finance Handling, Made Easy')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(400, 730, 'For queries, contact us at :budgettracker@gmail.com')
    c.setFillColor(colors.darkblue)
    c.setFont('Times-Roman', 20)
    c.drawString(250, 700, 'Monthly Report')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 670, 'Hi ' + u + ', ')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 650, "Here's your monthly report for the month of " + m2 + '.')
    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 15)
    c.drawString(40, 615, 'Account Information')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 590, 'Name: ' + name)
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 570, 'Email: ' + email)
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 550, 'Username : ' + u)
    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 15)
    c.drawString(40, 520, 'Expense for the month of ' + m2 + '.')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 12)
    c.drawString(220, 490, 'Table for Expenses and Budget')

    pdf_chart_colors = [
        HexColor('#a9cce3'),
        HexColor("#5dade2"),
        HexColor("#2980b9"),
        HexColor("#1f618d"),
        HexColor("#1b4f72"),
        HexColor("#154360"),
        HexColor("#fadbd8"),
        HexColor("#f1948a"),
        HexColor("#ec7063"),
        HexColor("#b03a2e"),
        HexColor("#943126"),
        HexColor('#d5f5e3')
    ]
    fn = (settings.MEDIA_ROOT + r'\uploads\Indiv_' + u + '_Data.xlsx')

    workbook = load_workbook(fn)
    ws = workbook['expenses']
    ws1 = workbook['budget']
    val = []
    val1 = []
    headers = []
    data = []
    for i in range(1, 13):
        ref = ws.cell(row=1, column=i)
        headers.append(ref.value.capitalize())

    headers[8] = 'Childcare'
    for i in range(1, 13):
        ref = ws.cell(row=m1 + 1, column=i)
        val1.append(ref.value)

    for i in range(1, 13):
        ref = ws1.cell(row=m1 + 1, column=i)
        val.append(ref.value)

    val.insert(0, 'Budget')
    val1.insert(0, 'Expenses')
    headers.insert(0, '      ')
    data.append(headers)
    data.append(val)
    data.append(val1)

    val2 = val1[:]
    del (val2[0])
    del (val2[-1])

    headers1 = headers[:]
    del (headers1[0])
    del (headers1[-1])
    data1 = []
    data1.append([])
    data1.append([])

    for i, x in enumerate(val2):
        data1[1].append(x)
    for i, x in enumerate(headers1):
        data1[0].append(x)

    total = []
    for j in range(2, 14):
        ref = ws.cell(row=j, column=12)
        ref_value = ref.value
        if ref_value == None:
            total.append(0)
        elif ref_value != None:
            ref_value = int(ref_value)
            total.append(ref_value)

    data2 = []
    data2.append(total)
    data2.append(
        [val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1]])

    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (1, 0), (12, 0), colors.mediumaquamarine),
        ('BACKGROUND', (0, 1), (0, 2), colors.mediumaquamarine),
        ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER')
    ]))

    t.wrapOn(c, 300, 300)
    t.drawOn(c, 20, 400)

    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 14)
    c.drawString(270, 360, 'Pie Chart')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(250, 345, 'Money Spent Over the Month')

    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(500, 100, 'Page 1 of 2')

    d = Drawing(300, 150)
    pc = Pie()
    pc.x = 100
    pc.y = 30
    pc.width = 150
    pc.height = 150
    pc.data = data1[1]
    pc.labels = data1[0]
    pc.slices.strokeColor = PCMYKColor(0, 0, 0, 0)
    pc.slices.label_visible = 0
    pc.slices.strokeColor = darkgray
    pc.slices.strokeWidth = 0.5

    d1 = Drawing(100, 100)
    legend = Legend()
    legend.dx = 5
    legend.dy = 5
    legend.fontName = 'Helvetica'
    legend.fontSize = 5.4
    legend.boxAnchor = 'w'
    legend.columnMaximum = 12
    legend.strokeWidth = 0.5
    legend.strokeColor = black
    legend.deltax = 75
    legend.deltay = 11
    legend.autoXPadding = 5
    legend.yGap = 2
    legend.dxTextSpace = 3
    legend.alignment = 'right'
    legend.dividerLines = 1 | 2 | 4
    legend.dividerOffsY = 5
    legend.subCols.rpad = 20

    def setItems(n, obj, attr, values):
        m = len(values)
        i = m // n
        for j in range(n):
            setattr(obj[j], attr, values[j * i % m])

    n = len(pc.data)
    setItems(n, pc.slices, 'fillColor', pdf_chart_colors)
    legend.colorNamePairs = [(pc.slices[i].fillColor, (pc.labels[i][0:20], '%0.2f' % pc.data[i])) for i in range(n)]

    d1.add(legend)
    d.add(pc)
    d.drawOn(c, 120, 120)
    d1.drawOn(c, 440, 220)

    c.showPage()
    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 15)
    c.drawString(250, 780, 'Bar Graph')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(190, 760, 'Comparison of expenses across various months with budget')
    d2 = Drawing(300, 300)

    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 50
    bc.valueAxis.valueMin = 0
    # bc.valueAxis.valueMax = 250000
    # bc.valueAxis.valueStep = 50000
    bc.height = 175
    bc.width = 450
    bc.categoryAxis.labels.dx = 8
    bc.categoryAxis.labels.dy = -2
    bc.data = data2
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.bars[(0, 0)].fillColor = colors.lightgreen
    bc.bars[(0, 1)].fillColor = colors.lightgreen
    bc.bars[(0, 2)].fillColor = colors.lightgreen
    bc.bars[(0, 3)].fillColor = colors.lightgreen
    bc.bars[(0, 4)].fillColor = colors.lightgreen
    bc.bars[(0, 5)].fillColor = colors.lightgreen
    bc.bars[(0, 6)].fillColor = colors.lightgreen
    bc.bars[(0, 7)].fillColor = colors.lightgreen
    bc.bars[(0, 8)].fillColor = colors.lightgreen
    bc.bars[(0, 9)].fillColor = colors.lightgreen
    bc.bars[(0, 10)].fillColor = colors.lightgreen
    bc.bars[(0, 11)].fillColor = colors.lightgreen

    bc.bars[(1, 0)].fillColor = colors.lightblue
    bc.bars[(1, 1)].fillColor = colors.lightblue
    bc.bars[(1, 2)].fillColor = colors.lightblue
    bc.bars[(1, 3)].fillColor = colors.lightblue
    bc.bars[(1, 4)].fillColor = colors.lightblue
    bc.bars[(1, 5)].fillColor = colors.lightblue
    bc.bars[(1, 6)].fillColor = colors.lightblue
    bc.bars[(1, 7)].fillColor = colors.lightblue
    bc.bars[(1, 8)].fillColor = colors.lightblue
    bc.bars[(1, 9)].fillColor = colors.lightblue
    bc.bars[(1, 10)].fillColor = colors.lightblue
    bc.bars[(1, 11)].fillColor = colors.lightblue

    bc.categoryAxis.categoryNames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    bc.categoryAxis.labels.angle = 30

    d2.add(bc)
    d2.drawOn(c, 40, 480)

    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 8.5)
    c.drawString(30, 430,
                 'Note : For any queries regarding incorrect info or wrong calculation, contact our customer care. We will get back to you shortly.')

    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(500, 100, 'Page 2 of 2')
    c.save()
    file_path = os.path.join(settings.MEDIA_ROOT, saving)

    try:
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/pdf")
                print(response)
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
    except:
        raise Http404

    finally:
        if os.path.isfile(saving):
            os.remove(saving)

    return render(request, 'accounts/indiv_pdf.html')

def corp_pdf_view(request):

    u = request.user.username
    user = Org_User.objects.get(username=u)
    f_name = user.c_name
    # l_name = user.l_name
    email = user.email
    name = f_name
    saving = settings.MEDIA_ROOT + r'\uploads\Report_for_' + u + '_General_Department.pdf'
    c = canvas.Canvas(saving, pagesize=A4)
    if m10 == 1:
        m2 = 'January'
    if m10 == 2:
        m2 = 'February'
    if m10 == 3:
        m2 = 'March'
    if m10 == 4:
        m2 = 'April'
    if m10 == 5:
        m2 = 'May'
    if m10 == 6:
        m2 = 'June'
    if m10 == 7:
        m2 = 'July'
    if m10 == 8:
        m2 = 'August'
    if m10 == 9:
        m2 = 'September'
    if m10 == 10:
        m2 = 'October'
    if m10 == 11:
        m2 = 'November'
    if m10 == 12:
        m2 = 'December'
    print(m10)
    print(m2)
    c.setFont('Times-Roman', 30)
    c.setFillColor(colors.deepskyblue)
    c.drawString(40, 780, 'Budget Tracker')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 13)
    c.drawString(55, 760, 'Finance Handling, Made Easy')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(400, 730, 'For queries, contact us at :budgettracker@gmail.com')
    c.setFillColor(colors.darkblue)
    c.setFont('Times-Roman', 20)
    c.drawString(250, 700, 'Monthly Report')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 670, 'Hi ' + u + ', ')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 650, "Here's your monthly report for the month of " + m2 + '.')
    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 15)
    c.drawString(40, 615, 'Account Information')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 590, 'Name: ' + name)
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 570, 'Email: ' + email)
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 9)
    c.drawString(40, 550, 'Username : ' + u)
    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 15)
    c.drawString(40, 520, 'Expense for the month of ' + m2 + '.')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 12)
    c.drawString(220, 490, 'Table for Expenses and Budget')

    pdf_chart_colors = [
        HexColor('#a9cce3'),
        HexColor("#5dade2"),
        HexColor("#2980b9"),
        HexColor("#1f618d"),
        HexColor("#1b4f72"),
        HexColor("#154360"),
        HexColor("#fadbd8"),
        HexColor("#f1948a"),
        HexColor("#ec7063"),
        HexColor("#b03a2e"),
        HexColor("#943126"),
        HexColor('#d5f5e3')
    ]
    fn = (settings.MEDIA_ROOT + r'\uploads\Organ_' + u + '_General.xlsx')

    workbook = load_workbook(fn)
    ws = workbook['expenses']
    ws1 = workbook['budget']
    val = []
    val1 = []
    headers = []
    data = []
    for i in range(1, 12):
        ref = ws.cell(row=1, column=i)
        headers.append(ref.value.capitalize())

    for i in range(1, 12):
        ref = ws.cell(row=m10 + 1, column=i)
        ref_value = ref.value
        if ref_value == None:
            ref_value = 0
        val1.append(ref_value)

    for i in range(1, 12):
        ref = ws1.cell(row=m10 + 1, column=i)
        ref_value = ref.value
        if ref_value == None:
            ref_value = 0
        val.append(ref_value)

    val.insert(0, 'Budget')
    val1.insert(0, 'Expenses')
    headers.insert(0, '      ')
    data.append(headers)
    data.append(val)
    data.append(val1)

    val2 = val1[:]
    del (val2[0])
    del (val2[-1])
    headers1 = headers[:]
    del (headers1[0])
    del (headers1[-1])
    data1 = []
    data1.append([])
    data1.append([])

    for i, x in enumerate(val2):
        data1[1].append(x)
    for i, x in enumerate(headers1):
        data1[0].append(x)

    total = []
    for j in range(2, 14):
        ref = ws.cell(row=j, column=11)
        ref_value = ref.value
        if ref_value == None:
            total.append(0)
        elif ref_value != None:
            ref_value = int(ref_value)
            total.append(ref_value)

    data2 = []
    data2.append(total)
    data2.append(
        [val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1], val[-1]])
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (1, 0), (11, 0), colors.mediumaquamarine),
        ('BACKGROUND', (0, 1), (0, 2), colors.mediumaquamarine),
        ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER')
    ]))

    t.wrapOn(c, 300, 300)
    t.drawOn(c, 20, 400)

    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 14)
    c.drawString(270, 360, 'Pie Chart')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(250, 345, 'Money Spent Over the Month')

    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(500, 100, 'Page 1 of 2')

    d = Drawing(300, 150)
    pc = Pie()
    pc.x = 100
    pc.y = 30
    pc.width = 150
    pc.height = 150
    pc.data = data1[1]
    pc.labels = data1[0]
    pc.slices.strokeColor = PCMYKColor(0, 0, 0, 0)
    pc.slices.label_visible = 0
    pc.slices.strokeColor = darkgray
    pc.slices.strokeWidth = 0.5

    d1 = Drawing(100, 100)
    legend = Legend()
    legend.dx = 5
    legend.dy = 5
    legend.fontName = 'Helvetica'
    legend.fontSize = 5.4
    legend.boxAnchor = 'w'
    legend.columnMaximum = 11
    legend.strokeWidth = 0.5
    legend.strokeColor = black
    legend.deltax = 75
    legend.deltay = 11
    legend.autoXPadding = 5
    legend.yGap = 2
    legend.dxTextSpace = 3
    legend.alignment = 'right'
    legend.dividerLines = 1 | 2 | 4
    legend.dividerOffsY = 5
    legend.subCols.rpad = 20

    def setItems(n, obj, attr, values):
        m = len(values)
        i = m // n
        for j in range(n):
            setattr(obj[j], attr, values[j * i % m])

    n = len(pc.data)
    setItems(n, pc.slices, 'fillColor', pdf_chart_colors)
    legend.colorNamePairs = [(pc.slices[i].fillColor, (pc.labels[i][0:20], '%0.2f' % pc.data[i])) for i in range(n)]

    d1.add(legend)
    d.add(pc)
    d.drawOn(c, 120, 120)
    d1.drawOn(c, 440, 220)

    c.showPage()
    c.setFillColor(colors.cornflowerblue)
    c.setFont('Times-Roman', 15)
    c.drawString(250, 780, 'Bar Graph')
    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(190, 760, 'Comparison of expenses across various months with budget')
    d2 = Drawing(300, 300)

    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 50
    bc.valueAxis.valueMin = 0
    # bc.valueAxis.valueMax = 250000
    # bc.valueAxis.valueStep = 50000
    bc.height = 175
    bc.width = 450
    bc.categoryAxis.labels.dx = 8
    bc.categoryAxis.labels.dy = -2
    bc.data = data2
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.bars[(0, 0)].fillColor = colors.lightgreen
    bc.bars[(0, 1)].fillColor = colors.lightgreen
    bc.bars[(0, 2)].fillColor = colors.lightgreen
    bc.bars[(0, 3)].fillColor = colors.lightgreen
    bc.bars[(0, 4)].fillColor = colors.lightgreen
    bc.bars[(0, 5)].fillColor = colors.lightgreen
    bc.bars[(0, 6)].fillColor = colors.lightgreen
    bc.bars[(0, 7)].fillColor = colors.lightgreen
    bc.bars[(0, 8)].fillColor = colors.lightgreen
    bc.bars[(0, 9)].fillColor = colors.lightgreen
    bc.bars[(0, 10)].fillColor = colors.lightgreen
    bc.bars[(0, 11)].fillColor = colors.lightgreen

    bc.bars[(1, 0)].fillColor = colors.lightblue
    bc.bars[(1, 1)].fillColor = colors.lightblue
    bc.bars[(1, 2)].fillColor = colors.lightblue
    bc.bars[(1, 3)].fillColor = colors.lightblue
    bc.bars[(1, 4)].fillColor = colors.lightblue
    bc.bars[(1, 5)].fillColor = colors.lightblue
    bc.bars[(1, 6)].fillColor = colors.lightblue
    bc.bars[(1, 7)].fillColor = colors.lightblue
    bc.bars[(1, 8)].fillColor = colors.lightblue
    bc.bars[(1, 9)].fillColor = colors.lightblue
    bc.bars[(1, 10)].fillColor = colors.lightblue
    bc.bars[(1, 11)].fillColor = colors.lightblue

    bc.categoryAxis.categoryNames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    bc.categoryAxis.labels.angle = 30

    d2.add(bc)
    d2.drawOn(c, 40, 480)

    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 8.5)
    c.drawString(30, 430,
                 'Note : For any queries regarding incorrect info or wrong calculation, contact our customer care. We will get back to you shortly.')

    c.setFillColor(colors.black)
    c.setFont('Times-Roman', 8)
    c.drawString(500, 100, 'Page 2 of 2')
    c.save()
    file_path = os.path.join(settings.MEDIA_ROOT, saving)
    try:
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/pdf")
                print(response)
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
    except:
        raise Http404

    finally:
        if os.path.isfile(saving):
            os.remove(saving)

    return render(request, 'accounts/corp_pdf.html')
